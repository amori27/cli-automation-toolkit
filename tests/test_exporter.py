from pathlib import Path
from openpyxl import load_workbook
from src.reporters.exporter import export_excel


class TestExportExcel:
    def test_creates_excel_file(self, tmp_path: Path) -> None:
        f = tmp_path / "in.csv"
        f.write_text("a,b\n1,2\n3,4\n")
        out = tmp_path / "report.xlsx"
        export_excel(str(f), str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.title == "Data"
        assert ws.cell(1, 1).value == "a"
        assert ws.cell(1, 2).value == "b"
        assert ws.cell(2, 1).value == 1
        assert ws.cell(3, 2).value == 4

    def test_excel_styling(self, tmp_path: Path) -> None:
        f = tmp_path / "in.csv"
        f.write_text("x\n1\n")
        out = tmp_path / "report.xlsx"
        export_excel(str(f), str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        header = ws.cell(1, 1)
        assert header.font.bold is True
        assert header.alignment.horizontal == 'center'
        assert ws.auto_filter.ref is not None

    def test_large_dataset(self, tmp_path: Path) -> None:
        f = tmp_path / "in.csv"
        import pandas as pd
        df = pd.DataFrame({"col": range(100)})
        df.to_csv(str(f), index=False)
        out = tmp_path / "report.xlsx"
        export_excel(str(f), str(out))
        wb = load_workbook(str(out))
        ws = wb.active
        assert ws.cell(101, 1).value == 99
